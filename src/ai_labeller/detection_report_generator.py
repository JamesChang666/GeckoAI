import base64
import io
import sys, re, os, json
import datetime as dt
from collections import defaultdict, Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


def _safe_detect_image_name(image_name):
    safe_name = str(image_name or "").strip() or "detect_result"
    safe_name = safe_name.replace("\\", "_").replace("/", "_").replace(":", "_")
    safe_name = safe_name.replace("*", "_").replace("?", "_").replace('"', "_")
    safe_name = safe_name.replace("<", "_").replace(">", "_").replace("|", "_")
    root, ext = os.path.splitext(safe_name)
    if not ext:
        ext = ".jpg"
    if not root:
        root = "detect_result"
    return root + ext


def _resolve_detect_image_path(csv_path, image_name, detect_image_path=""):
    direct = str(detect_image_path or "").strip()
    if direct and os.path.isfile(direct):
        return os.path.abspath(direct)
    detect_dir = os.path.join(os.path.dirname(os.path.abspath(csv_path)), "detected_images")
    if not os.path.isdir(detect_dir):
        return ""
    candidate = os.path.join(detect_dir, _safe_detect_image_name(image_name))
    if os.path.isfile(candidate):
        return os.path.abspath(candidate)
    # Backward-compatible fallback for renamed files (e.g. id/sub_id naming).
    safe = _safe_detect_image_name(image_name)
    safe_root, _safe_ext = os.path.splitext(safe)
    for name in os.listdir(detect_dir):
        p = os.path.join(detect_dir, name)
        if not os.path.isfile(p):
            continue
        root, _ext = os.path.splitext(name)
        if root == safe_root or root.startswith(safe_root + "_"):
            return os.path.abspath(p)
    return ""


def _record_id_key(row):
    raw_id = _valid_token(row.get("id", ""))
    raw_sub = _valid_token(row.get("sub_id", ""))
    name = str(row.get("image_name", "")).strip()
    board = _board_name_from_image_name(name)
    if raw_id and raw_sub:
        return f"{raw_id}/{raw_sub}"
    if raw_id:
        return raw_id
    if raw_sub:
        return f"{board}/{raw_sub}" if board else raw_sub
    if board:
        return board
    return name or "unknown"


def _parse_yolo_rects(label_path):
    out = []
    if not label_path or not os.path.isfile(label_path):
        return out
    try:
        with open(label_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception:
        return out
    for raw in lines:
        parts = raw.strip().split()
        if len(parts) < 5:
            continue
        cls = str(parts[0]).strip()
        if len(parts) >= 9:
            try:
                pts = list(map(float, parts[1:9]))
            except Exception:
                continue
            xs = [pts[0], pts[2], pts[4], pts[6]]
            ys = [pts[1], pts[3], pts[5], pts[7]]
            x1, x2 = min(xs), max(xs)
            y1, y2 = min(ys), max(ys)
        else:
            try:
                cx, cy, w, h = map(float, parts[1:5])
            except Exception:
                continue
            x1, y1 = cx - w / 2.0, cy - h / 2.0
            x2, y2 = cx + w / 2.0, cy + h / 2.0
        x1 = max(0.0, min(1.0, x1))
        y1 = max(0.0, min(1.0, y1))
        x2 = max(0.0, min(1.0, x2))
        y2 = max(0.0, min(1.0, y2))
        if x2 <= x1 or y2 <= y1:
            continue
        out.append((cls, (x1, y1, x2, y2)))
    return out


def _image_to_data_uri(path, max_side=1200):
    if not path or not os.path.isfile(path):
        return ""
    try:
        from PIL import Image

        img = Image.open(path).convert("RGB")
        w, h = img.size
        scale = min(1.0, float(max_side) / float(max(1, max(w, h))))
        if scale < 1.0:
            nw = max(1, int(round(w * scale)))
            nh = max(1, int(round(h * scale)))
            img = img.resize((nw, nh), getattr(Image, "Resampling", Image).LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88, optimize=True)
        data = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/jpeg;base64,{data}"
    except Exception:
        return ""


def _golden_labeled_data_uri(image_path, label_path, max_side=1200):
    if not image_path or not os.path.isfile(image_path):
        return ""
    try:
        from PIL import Image, ImageDraw

        img = Image.open(image_path).convert("RGB")
        w, h = img.size
        draw = ImageDraw.Draw(img)
        for cls_id, (x1, y1, x2, y2) in _parse_yolo_rects(label_path):
            px1 = int(round(x1 * w))
            py1 = int(round(y1 * h))
            px2 = int(round(x2 * w))
            py2 = int(round(y2 * h))
            draw.rectangle([(px1, py1), (px2, py2)], outline=(255, 99, 71), width=3)
            draw.text((px1 + 4, max(0, py1 - 16)), f"id:{cls_id}", fill=(255, 99, 71))
        scale = min(1.0, float(max_side) / float(max(1, max(w, h))))
        if scale < 1.0:
            nw = max(1, int(round(w * scale)))
            nh = max(1, int(round(h * scale)))
            img = img.resize((nw, nh), getattr(Image, "Resampling", Image).LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=90, optimize=True)
        data = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/jpeg;base64,{data}"
    except Exception:
        return _image_to_data_uri(image_path, max_side=max_side)

def load_data(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip().lstrip("\ufeff").lower().replace(" ", "_") for c in df.columns]

    has_golden = "status" in df.columns
    records = []
    for _, row in df.iterrows():
        name     = str(row.get("image_name", "")).strip()
        detected = str(row.get("detected_classes", "")).strip()
        classes  = parse_classes(detected)
        status   = str(row.get("status", "N/A")).strip() if has_golden else "N/A"
        details  = str(row.get("details", "")).strip() if has_golden else ""
        iou_val  = parse_iou_from_details(details)
        matched, total = parse_matched(details)
        records.append({
            "timestamp":        str(row.get("timestamp", "")).strip(),
            "image_name":       name,
            "id":               str(row.get("id", "")).strip(),
            "sub_id":           str(row.get("sub_id", "")).strip(),
            "prefix":           _record_id_key({"id": row.get("id", ""), "image_name": name}),
            "detected_classes": detected,
            "golden_mode":      str(row.get("golden_mode", "")).strip() if has_golden else "",
            "iou_threshold":    str(row.get("iou_threshold", "")).strip() if has_golden else "",
            "status":           status,
            "reason":           str(row.get("reason", "")).strip() if has_golden else "",
            "details":          details,
            "avg_iou":          iou_val,
            "matched":          matched,
            "total":            total,
            "total_components": sum(classes.values()),
            "num_classes":      len(classes),
            "classes_dict":     classes,
            "has_golden":       has_golden,
            "golden_image_path": str(row.get("golden_image_path", "")).strip() if has_golden else "",
            "golden_label_path": str(row.get("golden_label_path", "")).strip() if has_golden else "",
            "detect_image_path": _resolve_detect_image_path(
                path,
                name,
                str(row.get("detect_image_path", "")).strip(),
            ),
        })
    return records, has_golden


def _parse_report_timestamp(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return dt.datetime.strptime(text, fmt)
        except Exception:
            continue
    try:
        return dt.datetime.fromisoformat(text)
    except Exception:
        return None


def _compute_total_duration(records):
    stamps = [_parse_report_timestamp(r.get("timestamp", "")) for r in records]
    stamps = [s for s in stamps if s is not None]
    if len(stamps) < 2:
        return None
    return max(0, int(round((max(stamps) - min(stamps)).total_seconds())))


def _format_duration_text(total_seconds):
    if total_seconds is None:
        return "N/A"
    total_seconds = max(0, int(total_seconds))
    hours, rem = divmod(total_seconds, 3600)
    minutes, seconds = divmod(rem, 60)
    if hours > 0:
        return f"{hours}h {minutes}m {seconds}s"
    if minutes > 0:
        return f"{minutes}m {seconds}s"
    return f"{seconds}s"


def parse_classes(s):
    result = {}
    for item in s.split(";"):
        m = re.match(r"(.+)\s+x(\d+)", item.strip())
        if m:
            result[m.group(1).strip()] = int(m.group(2))
    return result


def parse_iou_from_details(d):
    m = re.search(r"avg IoU=([0-9.]+)", str(d))
    return float(m.group(1)) if m else None


def parse_matched(d):
    m = re.search(r"(\d+)/(\d+) matched", str(d))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def get_prefix(name):
    return str(name or "").strip() or "unknown"


def _valid_token(raw):
    text = str(raw or "").strip()
    bad = {"", "none", "null", "n/a", "na", "no_id", "unreadable_id", "no_sub_id", "unreadable_sub_id", "??"}
    return text if text.lower() not in bad else ""


def _board_name_from_image_name(image_name):
    name = str(image_name or "").strip()
    if "::" in name:
        name = name.split("::", 1)[0].strip()
    base = os.path.splitext(os.path.basename(name))[0]
    # Normalize piece/cut suffixes so reports can group by pre-cut board.
    base = re.sub(r"(_piece_\d+)$", "", base, flags=re.IGNORECASE)
    base = re.sub(r"(_cut_\d+(?:_\d+)?)$", "", base, flags=re.IGNORECASE)
    return base or ""


def _board_key(record):
    id_text = _valid_token(record.get("id", ""))
    if id_text:
        return id_text
    root = _board_name_from_image_name(record.get("image_name", ""))
    return root or "unknown"


def _is_piece_record(record):
    name = str(record.get("image_name", "")).strip().lower()
    sub_id = _valid_token(record.get("sub_id", ""))
    if sub_id:
        return True
    return ("_piece_" in name) or ("_cut_" in name)


def _piece_number_from_image_name(image_name):
    name = str(image_name or "").strip()
    if "::" in name:
        name = name.split("::", 1)[0].strip()
    base = os.path.splitext(os.path.basename(name))[0]
    m = re.search(r"_piece_(\d+)$", base, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"_cut_(\d+)(?:_\d+)?$", base, flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def build_supervisor_metrics(records):
    board_keys = set()
    board_fail = set()
    piece_total = 0
    piece_ok = 0
    piece_fail = 0
    part_total = defaultdict(int)
    part_fail = defaultdict(int)

    for r in records:
        bkey = _board_key(r)
        board_keys.add(bkey)
        status = str(r.get("status", "")).strip().upper()
        if status == "FAIL":
            board_fail.add(bkey)
        if _is_piece_record(r):
            piece_total += 1
            if status == "PASS":
                piece_ok += 1
            elif status == "FAIL":
                piece_fail += 1
        cls_dict = r.get("classes_dict", {}) or {}
        for cls, cnt in cls_dict.items():
            n = int(cnt)
            part_total[cls] += n
            if status == "FAIL":
                part_fail[cls] += n

    part_rows = []
    for cls in sorted(part_total.keys(), key=lambda k: part_total[k], reverse=True):
        total = int(part_total[cls])
        fail = int(part_fail.get(cls, 0))
        fail_rate = (fail / total * 100.0) if total > 0 else 0.0
        part_rows.append((cls, total, fail, fail_rate))

    board_total = len(board_keys)
    board_ng = len(board_fail)
    board_ok = max(0, board_total - board_ng)
    piece_ng_rate = (piece_fail / piece_total * 100.0) if piece_total > 0 else 0.0
    return {
        "board_total": board_total,
        "board_ok": board_ok,
        "board_ng": board_ng,
        "piece_total": piece_total,
        "piece_ok": piece_ok,
        "piece_fail": piece_fail,
        "piece_fail_rate": piece_ng_rate,
        "part_rows": part_rows,
    }


def aggregate(records):
    class_totals    = defaultdict(int)
    class_img_count = defaultdict(int)
    prefix_stats    = defaultdict(lambda: {
        "count": 0, "total_components": 0,
        "pass": 0, "fail": 0, "iou_sum": 0, "iou_count": 0
    })
    status_counts   = Counter()
    iou_values      = []

    for r in records:
        for cls, cnt in r["classes_dict"].items():
            class_totals[cls]    += cnt
            class_img_count[cls] += 1
        ps = prefix_stats[r["prefix"]]
        ps["count"]            += 1
        ps["total_components"] += r["total_components"]
        s = r["status"].upper()
        status_counts[s] += 1
        if s == "PASS": ps["pass"] += 1
        elif s == "FAIL": ps["fail"] += 1
        if r["avg_iou"] is not None:
            ps["iou_sum"]   += r["avg_iou"]
            ps["iou_count"] += 1
            iou_values.append(r["avg_iou"])

    sorted_classes = sorted(class_totals.items(), key=lambda x: x[1], reverse=True)
    return sorted_classes, class_img_count, prefix_stats, status_counts, iou_values

def build_excel(records, sorted_classes, class_img_count, prefix_stats,
                status_counts, iou_values, has_golden, out_path):

    H_FILL  = PatternFill("solid", start_color="1A3A5C")
    H_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    S_FILL  = PatternFill("solid", start_color="2E75B6")
    S_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    T_FILL  = PatternFill("solid", start_color="4472C4")
    T_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CELL_F  = Font(name="Calibri", size=10)
    ALT_F   = PatternFill("solid", start_color="EBF3FB")
    KPI_F   = PatternFill("solid", start_color="D6E8F7")
    FAIL_F  = PatternFill("solid", start_color="FDDEDE")
    FAIL_FT = Font(bold=True, color="C00000", name="Calibri", size=10)
    PASS_F  = PatternFill("solid", start_color="D8F5E4")
    PASS_FT = Font(bold=True, color="1A7A3C", name="Calibri", size=10)
    NA_F    = PatternFill("solid", start_color="F0F0F0")
    NA_FT   = Font(color="888888", name="Calibri", size=10)

    def thdr(ws, row, cols):
        for col, text in cols:
            c = ws.cell(row=row, column=col, value=text)
            c.font = T_FONT; c.fill = T_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")

    def shdr(ws, row, col, text, span=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = S_FONT; c.fill = S_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col+span-1)

    def status_style(ws, row, col, status):
        c = ws.cell(row=row, column=col)
        s = status.upper()
        if s == "FAIL":   c.fill, c.font = FAIL_F, FAIL_FT
        elif s == "PASS": c.fill, c.font = PASS_F, PASS_FT
        else:             c.fill, c.font = NA_F, NA_FT

    wb = Workbook()
    total_img  = len(records)
    total_comp = sum(r["total_components"] for r in records)
    fail_cnt   = status_counts.get("FAIL", 0)
    pass_cnt   = status_counts.get("PASS", 0)
    na_cnt     = total_img - fail_cnt - pass_cnt
    pass_rate  = round(pass_cnt / total_img * 100, 1) if total_img else 0
    total_duration = _compute_total_duration(records)
    total_duration_text = _format_duration_text(total_duration)

    ws = wb.active
    ws.title = "Raw Data"

    if has_golden:
        headers    = ["Timestamp", "Image Name", "Status", "Reason", "ID",
                      "Detected Classes", "IoU Mode", "IoU Threshold",
                      "Matched/Total", "Avg IoU", "Details", "Components", "# Classes"]
        col_widths = [20, 22, 9, 24, 12, 55, 10, 14, 14, 10, 30, 12, 10]
        STATUS_COL = 3
    else:
        headers    = ["Timestamp", "Image Name", "ID",
                      "Detected Classes", "Components", "# Classes"]
        col_widths = [20, 22, 12, 70, 12, 10]
        STATUS_COL = None

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    for i, r in enumerate(records, 2):
        row_fill = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        if has_golden:
            matched_str = f"{r['matched']}/{r['total']}" if r['matched'] is not None else "??"
            iou_str     = f"{r['avg_iou']:.3f}" if r['avg_iou'] is not None else "??"
            vals = [r["timestamp"], r["image_name"], r["status"], r.get("reason", ""), r["prefix"],
                    r["detected_classes"], r["golden_mode"], r["iou_threshold"],
                    matched_str, iou_str, r["details"],
                    r["total_components"], r["num_classes"]]
        else:
            vals = [r["timestamp"], r["image_name"], r["prefix"],
                    r["detected_classes"], r["total_components"], r["num_classes"]]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = CELL_F; cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == (6 if has_golden else 4)))

        if STATUS_COL:
            status_style(ws, i, STATUS_COL, r["status"])

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")

    shdr(ws2, 1, 1, "Detection Summary Report", 4)
    ws2.row_dimensions[1].height = 26

    kpis = [("Total Images", total_img), ("PASS", pass_cnt), ("FAIL", fail_cnt)]
    if has_golden:
        kpis.append(("Pass Rate", f"{pass_rate}%"))
        if iou_values:
            kpis.append(("Avg IoU", f"{round(sum(iou_values)/len(iou_values),3)}"))
    kpis.append(("Total Duration", total_duration_text))
    kpis.append(("Component Classes", len(sorted_classes)))

    ws2.cell(row=3, column=1, value="Metric").font = Font(bold=True, name="Calibri")
    ws2.cell(row=3, column=2, value="Value").font  = Font(bold=True, name="Calibri")
    for i, (metric, val) in enumerate(kpis, 4):
        ws2.cell(row=i, column=1, value=metric).font = CELL_F
        c = ws2.cell(row=i, column=2, value=val)
        c.font = Font(name="Calibri", size=11, bold=True); c.fill = KPI_F
        if metric == "PASS": c.font = Font(name="Calibri", size=11, bold=True, color="1A7A3C")
        if metric == "FAIL": c.font = Font(name="Calibri", size=11, bold=True, color="C00000")

    for col, w in zip(["A","B"], [28, 18]):
        ws2.column_dimensions[col].width = w

    sup = build_supervisor_metrics(records)
    sup_row = 4 + len(kpis) + 2
    shdr(ws2, sup_row, 1, "Production Summary", 4)
    thdr(ws2, sup_row + 1, [(1, "Metric"), (2, "Value"), (3, "Metric"), (4, "Value")])
    sup_items = [
        ("Boards (before cut)", sup["board_total"], "Board OK", sup["board_ok"]),
        ("Board NG", sup["board_ng"], "Pieces (after cut)", sup["piece_total"]),
        ("Piece OK", sup["piece_ok"], "Piece NG", sup["piece_fail"]),
        ("Piece NG Rate", f"{sup['piece_fail_rate']:.1f}%", "", ""),
    ]
    for i, (m1, v1, m2, v2) in enumerate(sup_items, sup_row + 2):
        fill = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        ws2.cell(row=i, column=1, value=m1).font = CELL_F
        ws2.cell(row=i, column=2, value=v1).font = CELL_F
        ws2.cell(row=i, column=3, value=m2).font = CELL_F
        ws2.cell(row=i, column=4, value=v2).font = CELL_F
        for col in (1, 2, 3, 4):
            ws2.cell(row=i, column=col).fill = fill

    part_row = sup_row + 7
    shdr(ws2, part_row, 1, "Part Failure Rate", 4)
    thdr(ws2, part_row + 1, [(1, "Part"), (2, "Total Qty"), (3, "NG Qty"), (4, "NG %")])
    for i, (part, total, fail, rate) in enumerate(sup["part_rows"], part_row + 2):
        fill = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        vals = [part, total, fail, f"{rate:.1f}%"]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font = CELL_F
            c.fill = fill

    start_row = part_row + 3 + len(sup["part_rows"])
    shdr(ws2, start_row, 1, "Results by ID", 5)
    if has_golden:
        thdr(ws2, start_row+1, [(1,"ID"),(2,"Images"),(3,"PASS"),(4,"FAIL"),(5,"Avg IoU")])
    else:
        thdr(ws2, start_row+1, [(1,"ID"),(2,"Images"),(3,"Total Components"),(4,"Avg Comp/Image"),(5,"")])

    for r_idx, (prefix, stats) in enumerate(sorted(prefix_stats.items()), start_row+2):
        fill = ALT_F if r_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        if has_golden:
            avg_iou = round(stats["iou_sum"]/stats["iou_count"], 3) if stats["iou_count"] else "??"
            vals = [prefix, stats["count"], stats["pass"], stats["fail"], avg_iou]
        else:
            avg_c = round(stats["total_components"]/stats["count"], 1) if stats["count"] else 0
            vals  = [prefix, stats["count"], stats["total_components"], avg_c, ""]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=r_idx, column=col, value=val)
            c.font = CELL_F; c.fill = fill
        if has_golden:
            p_c = ws2.cell(row=r_idx, column=3)
            p_c.font = Font(name="Calibri", size=10, bold=True, color="1A7A3C")
            f_c = ws2.cell(row=r_idx, column=4)
            f_c.font = Font(name="Calibri", size=10, bold=True, color="C00000")

    for col, w in zip(["C","D","E"], [10, 10, 12]):
        ws2.column_dimensions[col].width = w

    if has_golden and len(prefix_stats) > 0:
        id_last = start_row + 1 + len(prefix_stats)
        id_chart = BarChart()
        id_chart.type = "col"
        id_chart.title = "Images by ID (PASS/FAIL)"
        id_chart.style = 10
        id_chart.width = 18
        id_chart.height = 10
        id_chart.grouping = "clustered"
        id_chart.overlap = 0
        id_chart.add_data(
            Reference(ws2, min_col=3, max_col=4, min_row=start_row + 1, max_row=id_last),
            titles_from_data=True,
        )
        id_chart.set_categories(Reference(ws2, min_col=1, min_row=start_row + 2, max_row=id_last))
        try:
            id_chart.series[0].graphicalProperties.solidFill = "22C55E"
            id_chart.series[0].graphicalProperties.line.solidFill = "22C55E"
            id_chart.series[1].graphicalProperties.solidFill = "EF4444"
            id_chart.series[1].graphicalProperties.line.solidFill = "EF4444"
        except Exception:
            pass
        ws2.add_chart(id_chart, "G2")

    ws3 = wb.create_sheet("Class Analysis")
    if has_golden:
        shdr(ws3, 1, 1, "Component Class Detection Analysis (PASS/FAIL)", 6)
        thdr(
            ws3,
            2,
            [
                (1, "Class"),
                (2, "Total Detected"),
                (3, "Images Found In"),
                (4, "Avg per Image"),
                (5, "PASS Components"),
                (6, "FAIL Components"),
            ],
        )
    else:
        shdr(ws3, 1, 1, "Component Class Detection Analysis", 4)
        thdr(ws3, 2, [(1,"Class"),(2,"Total Detected"),(3,"Images Found In"),(4,"Avg per Image")])

    class_pass_totals = defaultdict(int)
    class_fail_totals = defaultdict(int)
    if has_golden:
        for r in records:
            status = str(r.get("status", "")).strip().upper()
            for cls, cnt in (r.get("classes_dict") or {}).items():
                if status == "PASS":
                    class_pass_totals[cls] += int(cnt)
                elif status == "FAIL":
                    class_fail_totals[cls] += int(cnt)

    for i, (cls, total) in enumerate(sorted_classes, 3):
        img_cnt = class_img_count[cls]
        avg     = round(total/img_cnt, 2) if img_cnt else 0
        fill    = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        if has_golden:
            vals = [cls, total, img_cnt, avg, class_pass_totals.get(cls, 0), class_fail_totals.get(cls, 0)]
        else:
            vals = [cls, total, img_cnt, avg]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.font = CELL_F; c.fill = fill
    if has_golden:
        for col, w in zip(["A","B","C","D","E","F"], [22, 16, 20, 16, 16, 16]):
            ws3.column_dimensions[col].width = w
    else:
        for col, w in zip(["A","B","C","D"], [22, 16, 20, 16]):
            ws3.column_dimensions[col].width = w

    chart = BarChart()
    chart.type = "col"; chart.title = "Components by Class"
    chart.style = 10; chart.width = 22; chart.height = 13
    last = 2 + len(sorted_classes)
    if has_golden:
        chart.grouping = "clustered"
        chart.overlap = 0
        chart.add_data(Reference(ws3, min_col=5, max_col=6, min_row=2, max_row=last), titles_from_data=True)
        try:
            chart.series[0].graphicalProperties.solidFill = "22C55E"
            chart.series[0].graphicalProperties.line.solidFill = "22C55E"
            chart.series[1].graphicalProperties.solidFill = "EF4444"
            chart.series[1].graphicalProperties.line.solidFill = "EF4444"
        except Exception:
            pass
    else:
        chart.add_data(Reference(ws3, min_col=2, min_row=2, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws3, min_col=1, min_row=3, max_row=last))
    ws3.add_chart(chart, "F2")

    if has_golden and iou_values:
        ws4 = wb.create_sheet("IoU Analysis")
        shdr(ws4, 1, 1, "IoU Score Analysis by ID", 4)
        thdr(ws4, 2, [(1,"ID"),(2,"Images"),(3,"PASS"),(4,"FAIL"),
                      (5,"Avg IoU"),(6,"Min IoU"),(7,"Max IoU")])

        cat_iou = defaultdict(list)
        for r in records:
            if r["avg_iou"] is not None:
                cat_iou[r["prefix"]].append((r["avg_iou"], r["status"]))

        for i, (cat, items) in enumerate(sorted(cat_iou.items()), 3):
            ious      = [x[0] for x in items]
            pass_n    = sum(1 for x in items if x[1].upper()=="PASS")
            fail_n    = sum(1 for x in items if x[1].upper()=="FAIL")
            fill      = ALT_F if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            vals      = [cat, len(items), pass_n, fail_n,
                         round(sum(ious)/len(ious),3), round(min(ious),3), round(max(ious),3)]
            for col, val in enumerate(vals, 1):
                c = ws4.cell(row=i, column=col, value=val)
                c.font = CELL_F; c.fill = fill

        for col, w in zip(["A","B","C","D","E","F","G"], [14,10,10,10,12,12,12]):
            ws4.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"  Excel -> {out_path}")

def build_html(records, sorted_classes, class_img_count, prefix_stats,
               status_counts, iou_values, has_golden, out_path):

    total_img  = len(records)
    fail_cnt   = status_counts.get("FAIL", 0)
    pass_cnt   = status_counts.get("PASS", 0)
    pass_rate  = round(pass_cnt/total_img*100, 1) if total_img else 0
    total_comp = sum(r["total_components"] for r in records)
    avg_comp   = round(total_comp/total_img, 1) if total_img else 0
    avg_iou    = round(sum(iou_values)/len(iou_values), 3) if iou_values else None
    ts_label   = str(records[0]["timestamp"]) if records else "N/A"
    total_duration = _compute_total_duration(records)
    total_duration_text = _format_duration_text(total_duration)
    sup = build_supervisor_metrics(records)

    cat_data   = sorted(prefix_stats.items())
    grand      = sum(t for _, t in sorted_classes) or 1

    cls_labels = json.dumps([c for c, _ in sorted_classes])
    cls_totals = json.dumps([t for _, t in sorted_classes])
    cls_imgs   = json.dumps([class_img_count[c] for c, _ in sorted_classes])
    cat_labels = json.dumps([c for c, _ in cat_data])
    cat_counts = json.dumps([s["count"] for _, s in cat_data])
    cat_comps  = json.dumps([s["total_components"] for _, s in cat_data])
    cat_pass   = json.dumps([s["pass"] for _, s in cat_data])
    cat_fail   = json.dumps([s["fail"] for _, s in cat_data])
    class_pass = []
    class_fail = []
    if has_golden:
        class_pass_map = defaultdict(int)
        class_fail_map = defaultdict(int)
        for r in records:
            status = str(r.get("status", "")).strip().upper()
            for cls, cnt in (r.get("classes_dict") or {}).items():
                if status == "PASS":
                    class_pass_map[cls] += int(cnt)
                elif status == "FAIL":
                    class_fail_map[cls] += int(cnt)
        class_pass = [class_pass_map.get(c, 0) for c, _ in sorted_classes]
        class_fail = [class_fail_map.get(c, 0) for c, _ in sorted_classes]

    iou_hist_labels, iou_hist_vals = [], []
    if has_golden and iou_values:
        bins = [(i/10, (i+1)/10) for i in range(10)]
        for lo, hi in bins:
            iou_hist_labels.append(f"{lo:.1f}-{hi:.1f}")
            iou_hist_vals.append(sum(1 for v in iou_values if lo <= v < hi))
        iou_hist_vals[-1] += sum(1 for v in iou_values if v == 1.0)

    palette = ["#3b82f6","#06b6d4","#8b5cf6","#ec4899","#f97316",
               "#10b981","#f59e0b","#60a5fa","#ef4444","#6366f1","#22c55e","#84cc16"]

    cls_rows = ""
    for i, (cls, total) in enumerate(sorted_classes):
        img_cnt = class_img_count[cls]
        avg     = round(total/img_cnt, 2) if img_cnt else 0
        pct     = round(total/grand*100, 1)
        bar_w   = round(total/(sorted_classes[0][1] or 1)*100)
        color   = palette[i % len(palette)]
        cls_rows += (f'<tr>'
                     f'<td><span class="dot" style="background:{color}"></span>{cls}</td>'
                     f'<td class="num">{total}</td><td class="num">{img_cnt}</td>'
                     f'<td class="num">{avg}</td>'
                     f'<td><span class="pct">{pct}%</span>'
                     f'<div class="bar-bg"><div class="bar-fg" style="width:{bar_w}%;background:{color}80"></div></div></td>'
                     f'</tr>\n')

    cat_rows = ""
    for prefix, stats in cat_data:
        avg = round(stats["total_components"]/stats["count"], 1) if stats["count"] else 0
        avg_iou_cat = f"{round(stats['iou_sum']/stats['iou_count'],3):.3f}" if stats["iou_count"] else "??"
        pass_badge = f'<span class="badge pass">{stats["pass"]} PASS</span>' if has_golden else ""
        fail_badge = f'<span class="badge fail">{stats["fail"]} FAIL</span>' if has_golden else ""
        iou_cell   = f'<td class="num">{avg_iou_cat}</td>' if has_golden else ""
        cat_rows  += (f'<tr>'
                      f'<td><span class="cat-tag">{prefix}</span></td>'
                      f'<td class="num">{stats["count"]}</td>'
                      f'<td class="num">{stats["total_components"]}</td>'
                      f'<td class="num">{avg}</td>'
                      f'{iou_cell}'
                      f'<td>{pass_badge} {fail_badge}</td>'
                      f'</tr>\n')

    sup_cards = (
        f'<div class="kpi green"><div class="val">{sup["board_total"]}</div><div class="lbl">Boards (Before Cut)</div></div>'
        f'<div class="kpi green"><div class="val">{sup["piece_total"]}</div><div class="lbl">Pieces (After Cut)</div></div>'
        f'<div class="kpi red"><div class="val">{sup["piece_fail"]}</div><div class="lbl">Piece NG</div></div>'
        f'<div class="kpi blue"><div class="val">{sup["piece_ok"]}</div><div class="lbl">Piece OK</div></div>'
        f'<div class="kpi yellow"><div class="val">{sup["piece_fail_rate"]:.1f}%</div><div class="lbl">Piece NG Rate</div></div>'
    )
    part_rows_html = ""
    for part, total, fail, rate in sup["part_rows"]:
        part_rows_html += (
            f"<tr><td>{part}</td><td class='num'>{total}</td><td class='num'>{fail}</td><td class='num'>{rate:.1f}%</td></tr>\n"
        )

    iou_th     = '<th>Avg IoU</th>' if has_golden else ''
    iou_chart  = ""
    iou_kpi    = ""
    pass_fail_chart = ""
    compare_section = ""
    warning_banner_html = ""
    piece_heatmap_html = ""

    if has_golden:
        iou_kpi = f'<div class="kpi blue"><div class="val">{avg_iou if avg_iou is not None else "??"}</div><div class="lbl">Avg IoU</div></div>'
        iou_chart = f"""
    <div class="card">
      <div class="card-title">IoU Score Distribution</div>
      <canvas id="iouHist" class="chart-canvas"></canvas>
    </div>"""
        pass_fail_chart = ""
        golden_image_path = ""
        golden_label_path = ""
        for r in records:
            if not golden_image_path:
                p = str(r.get("golden_image_path", "")).strip()
                if p and os.path.isfile(p):
                    golden_image_path = p
            if not golden_label_path:
                p = str(r.get("golden_label_path", "")).strip()
                if p and os.path.isfile(p):
                    golden_label_path = p
            if golden_image_path and golden_label_path:
                break
        golden_preview_uri = _golden_labeled_data_uri(golden_image_path, golden_label_path, max_side=1600)
        cmp_rows = ""
        for r in records:
            status = str(r.get("status", "")).strip().upper() or "N/A"
            if status != "FAIL":
                continue
            det_img_uri = _image_to_data_uri(str(r.get("detect_image_path", "")).strip(), max_side=1280)
            if not det_img_uri:
                continue
            status_cls = "status-pass" if status == "PASS" else ("status-fail" if status == "FAIL" else "status-na")
            img_name = str(r.get("image_name", "")).strip()
            details = str(r.get("details", "")).strip()
            reason_text = str(r.get("reason", "")).strip() or "n/a"
            golden_html = f'<img src="{golden_preview_uri}" loading="lazy" />' if golden_preview_uri else '<div class="cmp-empty">Golden sample not available</div>'
            details_html = f'<div class="cmp-detail">{details}</div>' if details else ""
            cmp_rows += (
                f'<div class="cmp-card">'
                f'<div class="cmp-head"><span class="cmp-name">{img_name}</span><span class="cmp-status {status_cls}">{status}</span></div>'
                f'<div class="cmp-reason"><strong>Reason:</strong> {reason_text}</div>'
                f'<div class="cmp-grid">'
                f'<div class="cmp-pane"><div class="cmp-title">Golden Sample (Labeled)</div>{golden_html}</div>'
                f'<div class="cmp-pane"><div class="cmp-title">Detected Result</div><img src="{det_img_uri}" loading="lazy" /></div>'
                f'</div>{details_html}</div>'
            )
        compare_section = (
            '<div class="card"><div class="card-title">Golden vs Detection Comparison (FAIL only)</div>'
            + (cmp_rows if cmp_rows else '<div class="cmp-empty">No FAIL images found for comparison.</div>')
            + '</div>'
        )

        piece_by_board: dict[str, dict[int, str]] = defaultdict(dict)
        board_fail_sets: dict[str, set[int]] = defaultdict(set)
        for r in records:
            if not _is_piece_record(r):
                continue
            pnum = _piece_number_from_image_name(r.get("image_name", ""))
            if pnum is None:
                continue
            board = _board_key(r)
            status = str(r.get("status", "")).strip().upper()
            prev = piece_by_board[board].get(pnum, "")
            if status == "FAIL" or prev != "FAIL":
                piece_by_board[board][pnum] = status if status in {"PASS", "FAIL"} else "N/A"
            if status == "FAIL":
                board_fail_sets[board].add(pnum)

        if piece_by_board:
            heat_cards = []
            for board, piece_map in sorted(piece_by_board.items(), key=lambda kv: str(kv[0])):
                if not piece_map:
                    continue
                max_piece = max(piece_map.keys())
                cells = []
                for idx in range(1, max_piece + 1):
                    st = piece_map.get(idx, "N/A")
                    cls = "pass" if st == "PASS" else ("fail" if st == "FAIL" else "nd-pass")
                    title = f"p{idx:03d} - {st}"
                    cells.append(f'<div class="pc {cls}" title="{title}">p{idx:03d}</div>')
                fail_n = sum(1 for v in piece_map.values() if v == "FAIL")
                pass_n = sum(1 for v in piece_map.values() if v == "PASS")
                heat_cards.append(
                    '<div class="card">'
                    f'<div class="card-title">{board} <small>{len(piece_map)} pieces</small></div>'
                    f'<div class="badge-row"><span class="badge pass">PASS {pass_n}</span> <span class="badge fail">FAIL {fail_n}</span></div>'
                    f'<div class="hm-wrap"><div class="hm-grid">{"".join(cells)}</div></div>'
                    '</div>'
                )
            piece_heatmap_html = (
                '<div class="section-label">Piece Heatmap</div>'
                '<div class="card"><div class="legend-row">'
                '<div class="li"><div class="ld" style="background:rgba(0,214,143,.45)"></div>PASS</div>'
                '<div class="li"><div class="ld" style="background:rgba(255,59,48,.75)"></div>FAIL</div>'
                '<div class="li"><div class="ld" style="background:rgba(74,158,255,.35)"></div>No Data</div>'
                '</div></div>'
                '<div class="row2">'
                + "".join(heat_cards)
                + '</div>'
            )

        if sup["piece_total"] > 0 and sup["piece_fail"] > 0:
            fail_positions = sorted({p for s in board_fail_sets.values() for p in s})
            fail_pos_text = ", ".join(f"p{n:03d}" for n in fail_positions) if fail_positions else "none"
            sets = [s for s in board_fail_sets.values() if s]
            same_positions = len(sets) >= 2 and all(s == sets[0] for s in sets[1:])
            if same_positions and fail_positions:
                tail_msg = "All boards share identical fail positions, strongly indicating a fixed-position missing-part issue. Check feeder/supply at these positions first."
            elif fail_positions:
                tail_msg = "Fail positions show repeated hotspots across boards. Check recurring positions and feeder/supply stability first."
            else:
                tail_msg = "This batch contains failed pieces. Check fixture, feeder/supply, and lighting stability."
            warning_banner_html = (
                '<div class="banner amber">'
                '<div class="banner-icon">⚠</div>'
                '<div class="banner-body">'
                f'<h3>Localized Failure: {sup["piece_fail"]} FAIL pieces ({sup["piece_fail_rate"]:.1f}%)</h3>'
                f'<p>Fail positions: {fail_pos_text}. {tail_msg}</p>'
                '</div></div>'
            )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Detection Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600;700&family=Inter:wght@400;500;600;700;900&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root {{
  --bg:      #0d0f14;
  --surface: #141720;
  --card:    #1c2030;
  --border:  #252a3a;
  --text:    #e8ecf4;
  --muted:   #7a849a;
  --pass:    #22c55e;
  --fail:    #ef4444;
  --blue:    #3b82f6;
  --amber:   #f59e0b;
  --purple:  #8b5cf6;
  --dim:     #4a5266;
}}
* {{ box-sizing:border-box; margin:0; padding:0 }}
body {{ font-family:'Inter',sans-serif; background:var(--bg); color:var(--text); min-height:100vh }}

header {{
  background: var(--surface);
  border-bottom: 1px solid var(--border);
  padding: 0;
}}
    .header-inner {{
      max-width:100%; margin:0 auto; padding:20px 24px;
      display:flex; align-items:center; gap:20px;
    }}
.header-icon {{
  width:48px; height:48px; border-radius:12px;
  background:linear-gradient(135deg,#3b82f6,#8b5cf6);
  display:flex; align-items:center; justify-content:center;
  font-size:13px; font-weight:700; font-family:'IBM Plex Mono',monospace; letter-spacing:.08em; flex-shrink:0;
}}
.header-text h1 {{ font-size:20px; font-weight:700; letter-spacing:-0.3px }}
.header-text p  {{ font-size:12px; color:var(--muted); margin-top:3px; font-family:'IBM Plex Mono',monospace }}
.mode-badge {{
  margin-left:auto; padding:6px 14px; border-radius:20px; font-size:11px;
  font-weight:600; font-family:'IBM Plex Mono',monospace; letter-spacing:0.5px;
  background: {'#1a3a1a; color:#22c55e; border:1px solid #22c55e40' if has_golden else '#1a1a3a; color:#8b5cf6; border:1px solid #8b5cf640'};
}}

    .container {{ max-width:100%; margin:0 auto; padding:24px 24px }}
.section-label {{
  font-family:'IBM Plex Mono',monospace; font-size:10px; font-weight:600; letter-spacing:.2em;
  text-transform:uppercase; color:var(--dim); margin:4px 0 12px 0;
  padding-bottom:8px; border-bottom:1px solid var(--border);
}}

/* KPI */
	.kpi-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); gap:14px; margin-bottom:24px }}
.kpi {{ background:var(--card); border:1px solid var(--border); border-radius:12px; padding:18px 14px; text-align:center; position:relative; overflow:hidden }}
.kpi::before {{ content:''; position:absolute; top:0; left:0; right:0; height:3px }}
.kpi.red::before   {{ background:#ef4444 }}
.kpi.green::before {{ background:#22c55e }}
.kpi.blue::before  {{ background:#3b82f6 }}
.kpi.purple::before {{ background:#8b5cf6 }}
.kpi.yellow::before {{ background:#f59e0b }}
.kpi .val {{ font-size:28px; font-weight:700; font-family:'IBM Plex Mono',monospace }}
.kpi.red .val    {{ color:#ef4444 }}
.kpi.green .val  {{ color:#22c55e }}
.kpi.blue .val   {{ color:#3b82f6 }}
.kpi.purple .val {{ color:#8b5cf6 }}
.kpi.yellow .val {{ color:#f59e0b }}
.kpi .lbl {{ font-size:10px; color:var(--muted); text-transform:uppercase; letter-spacing:1px; margin-top:4px; font-weight:500 }}

/* Cards */
	.row2 {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(420px,1fr)); gap:18px; margin-bottom:18px }}
	.row3 {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(320px,1fr)); gap:18px; margin-bottom:18px }}
	.card {{ background:var(--card); border:1px solid var(--border); border-radius:12px; padding:20px; margin-bottom:18px }}
.card-title {{ font-size:13px; font-weight:700; margin-bottom:14px; display:flex; align-items:center; gap:8px }}
.card-title small {{ font-family:'IBM Plex Mono',monospace; font-size:10px; color:var(--dim); font-weight:400 }}
.cmp-card {{ border:1px solid #243048; border-radius:10px; padding:12px; margin-bottom:12px; background:#141d30 }}
.cmp-head {{ display:flex; justify-content:space-between; align-items:center; gap:10px; margin-bottom:10px }}
.cmp-name {{ font-size:12px; font-family:'IBM Plex Mono',monospace; color:#cbd5e1; word-break:break-all }}
.cmp-status {{ padding:3px 8px; border-radius:999px; font-size:10px; font-weight:700; letter-spacing:.5px }}
.status-pass {{ color:#22c55e; border:1px solid #22c55e55; background:#052e16 }}
.status-fail {{ color:#ef4444; border:1px solid #ef444455; background:#450a0a }}
.status-na {{ color:#94a3b8; border:1px solid #334155; background:#0f172a }}
.cmp-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:10px }}
.cmp-pane {{ border:1px solid #2a364f; border-radius:8px; background:#0f172a; padding:8px }}
.cmp-title {{ font-size:10px; color:#94a3b8; text-transform:uppercase; letter-spacing:1px; margin-bottom:6px }}
.cmp-pane img {{ width:100%; border-radius:6px; border:1px solid #1e293b; display:block }}
	.cmp-empty {{ color:#64748b; font-size:12px; padding:16px 8px }}
	.cmp-detail {{ margin-top:8px; color:#94a3b8; font-size:11px; line-height:1.5 }}
	.cmp-reason {{ margin-bottom:8px; color:#f59e0b; font-size:11px; line-height:1.4 }}

/* Tables */
.tbl-wrap {{ overflow-x:auto }}
table {{ width:100%; border-collapse:collapse; font-size:12px }}
th {{ background:#0d1526; color:var(--muted); font-size:10px; font-weight:600; text-transform:uppercase; letter-spacing:1px; padding:9px 12px; text-align:left; border-bottom:1px solid var(--border); white-space:nowrap }}
td {{ padding:8px 12px; border-bottom:1px solid #1a2030; color:#cbd5e1 }}
tr:last-child td {{ border-bottom:none }}
tr:hover td {{ background:#1f2d44 }}
td.num {{ font-family:'IBM Plex Mono',monospace; font-size:11px; color:#94a3b8 }}

/* Badges */
	.badge {{ display:inline-block; padding:3px 8px; border-radius:6px; font-size:10px; font-weight:600; font-family:'IBM Plex Mono',monospace; margin:1px }}
	.badge.pass {{ background:#052e16; color:#22c55e; border:1px solid #22c55e30 }}
	.badge.fail {{ background:#450a0a; color:#ef4444; border:1px solid #ef444430 }}
	.badge-row {{ margin-bottom:8px }}
	.cat-tag {{ display:inline-block; padding:3px 9px; border-radius:6px; font-size:10px; font-weight:600; background:#162038; color:#60a5fa; border:1px solid #3b82f630 }}
	.dot {{ display:inline-block; width:8px; height:8px; border-radius:50%; margin-right:8px; vertical-align:middle }}
	.pct {{ font-family:'IBM Plex Mono',monospace; font-size:10px; color:var(--muted) }}
	.bar-bg {{ height:4px; background:var(--border); border-radius:2px; margin-top:4px }}
	.bar-fg {{ height:100%; border-radius:2px }}
	.banner {{
	  border-radius:8px; padding:14px 18px; margin-bottom:18px;
	  display:flex; align-items:flex-start; gap:12px;
	  border-left-width:4px; border-left-style:solid; border:1px solid;
	}}
	.banner.amber {{ background:rgba(255,184,0,.08); border-color:rgba(255,184,0,.35); border-left-color:#ffb800 }}
	.banner-icon {{ font-size:18px; flex-shrink:0; margin-top:1px }}
	.banner-body h3 {{ font-size:13px; font-weight:700; margin-bottom:5px; color:#ffb800 }}
	.banner-body p {{ font-size:12px; color:var(--muted); line-height:1.65 }}
	.hm-wrap {{ margin-bottom:8px }}
	.hm-grid {{ display:flex; flex-wrap:wrap; gap:4px }}
	.pc {{
	  width:28px; height:28px; border-radius:4px; display:flex; align-items:center; justify-content:center;
	  font-family:'IBM Plex Mono',monospace; font-size:9px; font-weight:600; cursor:default; transition:transform .15s;
	  flex-shrink:0;
	}}
	.pc:hover {{ transform:scale(1.22); z-index:10; position:relative }}
	.pc.pass {{ background:rgba(0,214,143,.45); color:rgba(0,0,0,.75) }}
	.pc.fail {{ background:rgba(255,59,48,.75); color:#fff }}
	.pc.nd-pass {{ background:rgba(74,158,255,.35); color:#fff }}
	.legend-row {{ display:flex; gap:16px; flex-wrap:wrap }}
	.li {{ display:flex; align-items:center; gap:5px; font-size:11px; color:var(--muted) }}
	.ld {{ width:10px; height:10px; border-radius:2px; flex-shrink:0 }}
	.chart-canvas {{ width:100% !important; height:clamp(220px, 36vh, 420px) !important; display:block }}

/* Chart.js defaults */
	@media(max-width:1200px) {{
	  .container,.header-inner {{ padding:16px }}
	  .row2 {{ grid-template-columns:1fr }}
	}}
	@media(max-width:768px) {{
	  .row2,.row3 {{ grid-template-columns:1fr }}
	  .cmp-grid {{ grid-template-columns:1fr }}
	  .chart-canvas {{ height:clamp(200px, 42vh, 320px) !important }}
	  .kpi .val {{ font-size:22px }}
	}}
</style>
</head>
<body>
<header>
  <div class="header-inner">
    <div class="header-icon">gekoai</div>
    <div class="header-text">
      <h1>Component Detection Dashboard</h1>
      <p>{ts_label} &nbsp;|&nbsp; {total_img} images &nbsp;|&nbsp; Total duration: {total_duration_text} &nbsp;|&nbsp; {'IoU threshold: ' + str(records[0]['iou_threshold']) if has_golden else 'Detection-only mode'}</p>
    </div>
    <div class="mode-badge">{'GOLDEN COMPARISON' if has_golden else 'DETECTION ONLY'}</div>
  </div>
</header>
<div class="container">

  <!-- KPIs -->
  <div class="section-label">Overview</div>
  <div class="kpi-row">
    <div class="kpi blue"><div class="val">{total_img}</div><div class="lbl">Total Images</div></div>
    {'<div class="kpi green"><div class="val">' + str(pass_cnt) + '</div><div class="lbl">PASS</div></div>' if has_golden else ''}
    {'<div class="kpi red"><div class="val">' + str(fail_cnt) + '</div><div class="lbl">FAIL</div></div>' if has_golden else ''}
    {'<div class="kpi green"><div class="val">' + str(pass_rate) + '%</div><div class="lbl">Pass Rate</div></div>' if has_golden else ''}
    {iou_kpi}
    <div class="kpi blue"><div class="val">{total_duration_text}</div><div class="lbl">Total Duration</div></div>
    <div class="kpi purple"><div class="val">{len(sorted_classes)}</div><div class="lbl">Classes</div></div>
    <div class="kpi yellow"><div class="val">{total_comp}</div><div class="lbl">Components</div></div>
    <div class="kpi blue"><div class="val">{avg_comp}</div><div class="lbl">Avg/Image</div></div>
  </div>

  <div class="kpi-row">
    {sup_cards}
  </div>
  {warning_banner_html}
  {piece_heatmap_html}

  <!-- Charts row 1 -->
  <div class="section-label">Pass/Fail Overview</div>
  <div class="row2">
    <div class="card"><div class="card-title">Components by Class {'(PASS/FAIL)' if has_golden else ''}</div><canvas id="classBar" class="chart-canvas"></canvas></div>
    <div class="card"><div class="card-title">Images by ID {'(PASS/FAIL)' if has_golden else ''}</div><canvas id="catPie" class="chart-canvas"></canvas></div>
    {pass_fail_chart if has_golden else ''}
  </div>

  <!-- Charts row 2 -->
  <div class="section-label">Distribution</div>
  <div class="row2">
    <div class="card"><div class="card-title">Components per ID</div><canvas id="catComp" class="chart-canvas"></canvas></div>
    {iou_chart if has_golden else '<div class="card"><div class="card-title">Class Frequency <small># images</small></div><canvas id="clsFreq" class="chart-canvas"></canvas></div>'}
  </div>

  <!-- Class breakdown table -->
  <div class="section-label">Class Details</div>
  <div class="card">
    <div class="card-title">Component Class Breakdown</div>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>Class</th><th>Total</th><th>Images</th><th>Avg/Image</th><th>Share</th></tr></thead>
        <tbody>{cls_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ID table -->
  <div class="section-label">ID Details</div>
  <div class="card">
    <div class="card-title">ID Summary</div>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>ID</th><th>Images</th><th>Components</th><th>Avg/Image</th>{iou_th}<th>Results</th></tr></thead>
        <tbody>{cat_rows}</tbody>
      </table>
    </div>
  </div>
  <div class="card">
    <div class="card-title">Part-Level NG Rate</div>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>Part</th><th>Total Qty</th><th>NG Qty</th><th>NG %</th></tr></thead>
        <tbody>{part_rows_html if part_rows_html else '<tr><td colspan="4">No detected parts</td></tr>'}</tbody>
      </table>
    </div>
  </div>
  {compare_section}

</div>
<script>
const P  = ["#3b82f6","#06b6d4","#8b5cf6","#ec4899","#f97316","#10b981","#f59e0b","#60a5fa","#ef4444","#6366f1","#22c55e","#84cc16"];
const al = (cols, a) => cols.map(c => c + (a||'99'));
const g  = {{x:{{ticks:{{color:"#64748b",font:{{size:10}}}},grid:{{color:"#1a2540"}}}},y:{{ticks:{{color:"#64748b",font:{{size:10}}}},grid:{{color:"#243048"}}}}}};
const gPct = {{x:{{stacked:true,ticks:{{color:"#64748b",font:{{size:10}}}},grid:{{color:"#1a2540"}}}},y:{{stacked:true,min:0,max:100,ticks:{{color:"#64748b",font:{{size:10}},callback:(v)=>v+"%"}},grid:{{color:"#243048"}}}}}};
const lg = {{labels:{{color:"#94a3b8",font:{{size:11}}}}}};

const clsL = {cls_labels}, clsT = {cls_totals}, clsI = {cls_imgs};
const catL = {cat_labels}, catC = {cat_counts}, catM = {cat_comps};
const catP = {cat_pass}, catF = {cat_fail};
const clsP = {json.dumps(class_pass)}, clsF = {json.dumps(class_fail)};
const toPctPair = (passArr, failArr) => {{
  const p = [], f = [];
  for (let i = 0; i < Math.max(passArr.length, failArr.length); i++) {{
    const pv = Number(passArr[i] || 0), fv = Number(failArr[i] || 0);
    const t = pv + fv;
    p.push(t > 0 ? +(pv * 100 / t).toFixed(2) : 0);
    f.push(t > 0 ? +(fv * 100 / t).toFixed(2) : 0);
  }}
  return {{pass:p, fail:f}};
}};
const pfClass = toPctPair(clsP, clsF);
const pfId = toPctPair(catP, catF);
const ttPct = (rawP, rawF) => ({{
  callbacks: {{
    label: (ctx) => {{
      const i = ctx.dataIndex;
      const isPass = String(ctx.dataset.label || '').toUpperCase() === 'PASS';
      const raw = isPass ? Number(rawP[i] || 0) : Number(rawF[i] || 0);
      const pct = Number(ctx.parsed.y || 0);
      return `${{ctx.dataset.label}}: ${{raw}} (${{pct.toFixed(2)}}%)`;
    }}
  }}
}});

new Chart(document.getElementById("classBar"),{{
  type:"bar",
  data:{{labels:clsL,datasets:[{'{label:"PASS",data:pfClass.pass,backgroundColor:"#22c55e66",borderColor:"#22c55e",borderWidth:1,borderRadius:4,categoryPercentage:0.9,barPercentage:0.95},{label:"FAIL",data:pfClass.fail,backgroundColor:"#ef444466",borderColor:"#ef4444",borderWidth:1,borderRadius:4,categoryPercentage:0.9,barPercentage:0.95}' if has_golden else '{label:"Total",data:clsT,backgroundColor:al(P),borderColor:P,borderWidth:1,borderRadius:4}' }]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{'lg' if has_golden else '{display:false}'}, {'tooltip: ttPct(clsP, clsF),' if has_golden else ''} }},scales:{'gPct' if has_golden else 'g'}}}
}});

new Chart(document.getElementById("catPie"),{{
  type:{'"bar"' if has_golden else '"doughnut"'},
  data:{{labels:catL,datasets:[{'{label:"PASS",data:pfId.pass,backgroundColor:"#22c55e66",borderColor:"#22c55e",borderWidth:1,borderRadius:3,categoryPercentage:0.9,barPercentage:0.95},{label:"FAIL",data:pfId.fail,backgroundColor:"#ef444466",borderColor:"#ef4444",borderWidth:1,borderRadius:3,categoryPercentage:0.9,barPercentage:0.95}' if has_golden else '{data:catC,backgroundColor:al(P.slice(0,catL.length),\'cc\'),borderColor:P,borderWidth:2}' }]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:lg {' ,tooltip: ttPct(catP, catF)' if has_golden else ''} }}{' ,scales:gPct' if has_golden else ''}}}
}});

new Chart(document.getElementById("catComp"),{{
  type:"bar",
  data:{{labels:catL,datasets:[{{label:"Components",data:catM,backgroundColor:al(P.slice(0,catL.length),'aa'),borderColor:P,borderWidth:1,borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:g}}
}});

{'new Chart(document.getElementById("iouHist"),{type:"bar",data:{labels:' + json.dumps(iou_hist_labels) + ',datasets:[{label:"Images",data:' + json.dumps(iou_hist_vals) + ',backgroundColor:"#3b82f666",borderColor:"#3b82f6",borderWidth:1,borderRadius:3}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:g}});' if has_golden and iou_values else ''}

{'new Chart(document.getElementById("clsFreq"),{type:"bar",data:{labels:clsL,datasets:[{label:"Images",data:clsI,backgroundColor:"#8b5cf666",borderColor:"#8b5cf6",borderWidth:1,borderRadius:3}]},options:{responsive:true,maintainAspectRatio:false,indexAxis:"y",plugins:{legend:{display:false}},scales:g}});' if not has_golden else ''}
</script>
</body>
</html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML  -> {out_path}")

def build_pdf(records, sorted_classes, class_img_count, prefix_stats,
              status_counts, iou_values, has_golden, out_path):
    # Prefer generating PDF directly from the HTML dashboard so layout matches 1:1.
    html_path = os.path.splitext(out_path)[0] + ".html"
    if os.path.isfile(html_path):
        try:
            from weasyprint import HTML  # type: ignore

            HTML(filename=html_path).write_pdf(out_path)
            print(f"  PDF   -> {out_path} (from HTML)")
            return
        except Exception:
            pass
        try:
            import subprocess
            from pathlib import Path

            html_uri = Path(html_path).resolve().as_uri()
            browser_cmds = [
                [
                    "msedge",
                    "--headless=new",
                    "--disable-gpu",
                    "--virtual-time-budget=8000",
                    f"--print-to-pdf={out_path}",
                    html_uri,
                ],
                [
                    "chrome",
                    "--headless=new",
                    "--disable-gpu",
                    "--virtual-time-budget=8000",
                    f"--print-to-pdf={out_path}",
                    html_uri,
                ],
                [
                    "chromium",
                    "--headless=new",
                    "--disable-gpu",
                    "--virtual-time-budget=8000",
                    f"--print-to-pdf={out_path}",
                    html_uri,
                ],
            ]
            for cmd in browser_cmds:
                try:
                    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    if os.path.isfile(out_path) and os.path.getsize(out_path) > 0:
                        print(f"  PDF   -> {out_path} (from HTML)")
                        return
                except Exception:
                    continue
        except Exception:
            pass
        try:
            import subprocess

            subprocess.run(["wkhtmltopdf", html_path, out_path], check=True)
            print(f"  PDF   -> {out_path} (from HTML)")
            return
        except Exception:
            pass

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, PageBreak)
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie

    W, H  = A4
    DKBL  = colors.HexColor("#1A3A5C")
    MDBL  = colors.HexColor("#2E75B6")
    LTBL  = colors.HexColor("#D6E8F7")
    ALTBL = colors.HexColor("#EBF3FB")
    PASS  = colors.HexColor("#1A7A3C")
    PASS_BG = colors.HexColor("#D8F5E4")
    FAIL  = colors.HexColor("#C00000")
    FAIL_BG = colors.HexColor("#FDDEDE")
    NA    = colors.HexColor("#888888")
    WHITE = colors.white
    GRAY  = colors.HexColor("#64748b")
    PAL   = [colors.HexColor(h) for h in
             ["#3b82f6","#06b6d4","#8b5cf6","#ec4899","#f97316",
              "#10b981","#f59e0b","#60a5fa","#ef4444","#6366f1","#22c55e"]]

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=1.4*cm, rightMargin=1.4*cm,
                            topMargin=1.4*cm, bottomMargin=1.4*cm)

    body_style = ParagraphStyle("body", fontSize=8, fontName="Helvetica",
                                textColor=colors.HexColor("#1e293b"))

    total_img  = len(records)
    fail_cnt   = status_counts.get("FAIL", 0)
    pass_cnt   = status_counts.get("PASS", 0)
    pass_rate  = round(pass_cnt/total_img*100, 1) if total_img else 0
    total_comp = sum(r["total_components"] for r in records)
    avg_iou    = round(sum(iou_values)/len(iou_values), 3) if iou_values else None
    ts_label   = str(records[0]["timestamp"]) if records else ""
    total_duration = _compute_total_duration(records)
    total_duration_text = _format_duration_text(total_duration)
    cat_data   = sorted(prefix_stats.items())
    PW         = W - 2.8*cm  # printable width

    def sec_bar(text, sub=""):
        d = Drawing(PW, 24)
        d.add(Rect(0, 0, PW, 24, fillColor=MDBL, strokeColor=None))
        d.add(String(8, 7, text, fontSize=10, fillColor=WHITE, fontName="Helvetica-Bold"))
        if sub:
            d.add(String(PW - 6, 7, sub, fontSize=8, fillColor=colors.HexColor("#93c5fd"),
                         fontName="Helvetica", textAnchor="end"))
        return d

    def mk_table(hdr, rows, cws, status_col=None):
        ts = TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), MDBL),
            ("TEXTCOLOR",   (0,0), (-1,0), WHITE),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [WHITE, ALTBL]),
            ("GRID",        (0,0), (-1,-1), 0.25, colors.HexColor("#c8d8ec")),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ])
        t = Table([hdr]+rows, colWidths=cws, repeatRows=1)
        t.setStyle(ts)
        if status_col is not None:
            for ri, row in enumerate(rows, 1):
                s = str(row[status_col]).upper()
                if s == "PASS":
                    t.setStyle(TableStyle([
                        ("BACKGROUND",(status_col,ri),(status_col,ri), PASS_BG),
                        ("TEXTCOLOR", (status_col,ri),(status_col,ri), PASS),
                        ("FONTNAME",  (status_col,ri),(status_col,ri), "Helvetica-Bold"),
                    ]))
                elif s == "FAIL":
                    t.setStyle(TableStyle([
                        ("BACKGROUND",(status_col,ri),(status_col,ri), FAIL_BG),
                        ("TEXTCOLOR", (status_col,ri),(status_col,ri), FAIL),
                        ("FONTNAME",  (status_col,ri),(status_col,ri), "Helvetica-Bold"),
                    ]))
        return t

    story = []

    cov = Drawing(PW, 70)
    cov.add(Rect(0, 0, PW, 70, fillColor=DKBL, strokeColor=None, rx=4, ry=4))
    mode_txt = "GOLDEN COMPARISON MODE" if has_golden else "DETECTION-ONLY MODE"
    cov.add(String(14, 44, "Component Detection Report", fontSize=20,
                   fillColor=WHITE, fontName="Helvetica-Bold"))
    cov.add(String(14, 26, f"{ts_label}   繚   {total_img} images   繚   {total_duration_text}   繚   {mode_txt}",
                   fontSize=9, fillColor=colors.HexColor("#93c5fd"), fontName="Helvetica"))
    story.append(cov)
    story.append(Spacer(1, 14))

    story.append(sec_bar("Overview"))
    story.append(Spacer(1, 8))

    kpi_items = [("Total Images", str(total_img), DKBL)]
    if has_golden:
        kpi_items += [
            ("PASS",      str(pass_cnt),      PASS),
            ("FAIL",      str(fail_cnt),      FAIL),
            ("Pass Rate", f"{pass_rate}%",    MDBL),
        ]
        if avg_iou: kpi_items.append(("Avg IoU", str(avg_iou), colors.HexColor("#0e7490")))
    kpi_items += [
        ("Total Duration", total_duration_text, colors.HexColor("#1d4ed8")),
        ("Classes",    str(len(sorted_classes)), colors.HexColor("#6d28d9")),
        ("Components", str(total_comp),           colors.HexColor("#b45309")),
    ]

    cols_n = min(4, len(kpi_items))
    rows_n = -(-len(kpi_items) // cols_n)
    cw = PW / cols_n
    ch = 46
    kd = Drawing(PW, rows_n * (ch + 4))
    for idx, (lbl, val, col) in enumerate(kpi_items):
        ci = idx % cols_n
        ri = idx // cols_n
        x  = ci * cw
        y  = (rows_n - 1 - ri) * (ch + 4)
        kd.add(Rect(x+1, y+1, cw-3, ch-2, fillColor=LTBL,
                    strokeColor=MDBL, strokeWidth=0.4, rx=3, ry=3))
        kd.add(String(x+cw/2, y+ch/2+3, val, fontSize=17, fillColor=col,
                       fontName="Helvetica-Bold", textAnchor="middle"))
        kd.add(String(x+cw/2, y+6, lbl, fontSize=7, fillColor=GRAY,
                       fontName="Helvetica", textAnchor="middle"))
    story.append(kd)
    story.append(Spacer(1, 16))

    cls_names = [c for c, _ in sorted_classes]
    cls_vals  = [t for _, t in sorted_classes]
    class_pass_vals = []
    class_fail_vals = []
    if has_golden:
        class_pass_map = defaultdict(int)
        class_fail_map = defaultdict(int)
        for r in records:
            status = str(r.get("status", "")).strip().upper()
            for cls, cnt in (r.get("classes_dict") or {}).items():
                if status == "PASS":
                    class_pass_map[cls] += int(cnt)
                elif status == "FAIL":
                    class_fail_map[cls] += int(cnt)
        class_pass_vals = [class_pass_map.get(c, 0) for c in cls_names]
        class_fail_vals = [class_fail_map.get(c, 0) for c in cls_names]
    if cls_vals:
        story.append(sec_bar("Components by Class (PASS/FAIL)" if has_golden else "Total Detected Components by Class"))
        story.append(Spacer(1, 8))
        bc = VerticalBarChart()
        bc.x, bc.y = 55, 20
        bc.width   = PW - 70
        bc.height  = 130
        if has_golden:
            bc.data = [class_pass_vals, class_fail_vals]
        else:
            bc.data = [cls_vals]
        bc.categoryAxis.categoryNames = cls_names
        bc.categoryAxis.labels.angle  = 28
        bc.categoryAxis.labels.fontSize = 7
        bc.categoryAxis.labels.dy = -10
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 7
        if has_golden:
            bc.bars[0].fillColor = PASS
            bc.bars[1].fillColor = FAIL
        else:
            bc.bars[0].fillColor = MDBL
        d1 = Drawing(PW, 170)
        d1.add(bc)
        story.append(d1)
        story.append(Spacer(1, 12))

    story.append(sec_bar("Images by ID (PASS/FAIL)" if has_golden else "Image Distribution by ID"))
    story.append(Spacer(1, 8))
    if has_golden:
        id_names = [lbl for lbl, _ in cat_data]
        id_pass = [stats.get("pass", 0) for _, stats in cat_data]
        id_fail = [stats.get("fail", 0) for _, stats in cat_data]
        id_bar = VerticalBarChart()
        id_bar.x, id_bar.y = 55, 20
        id_bar.width = PW - 70
        id_bar.height = 130
        id_bar.data = [id_pass, id_fail]
        id_bar.categoryAxis.categoryNames = id_names
        id_bar.categoryAxis.labels.angle = 28
        id_bar.categoryAxis.labels.fontSize = 7
        id_bar.categoryAxis.labels.dy = -10
        id_bar.valueAxis.valueMin = 0
        id_bar.valueAxis.labels.fontSize = 7
        id_bar.bars[0].fillColor = PASS
        id_bar.bars[1].fillColor = FAIL
        id_d = Drawing(PW, 170)
        id_d.add(id_bar)
        story.append(id_d)
    else:
        pie_size   = 150
        total_cats = sum(s["count"] for _, s in cat_data) or 1
        pie_h      = pie_size + 30
        pie_d      = Drawing(PW, pie_h)
        pie        = Pie()
        pie.x      = 15
        pie.y      = 10
        pie.width  = pie.height = pie_size
        pie.data   = [s["count"] for _, s in cat_data]
        pie.labels = [f"{lbl}\n{round(s['count']/total_cats*100,1)}%" for lbl, s in cat_data]
        pie.sideLabels       = True
        pie.sideLabelsOffset = 0.08
        pie.simpleLabels     = False
        for i in range(len(cat_data)):
            pie.slices[i].fillColor   = PAL[i % len(PAL)]
            pie.slices[i].strokeWidth = 1
            pie.slices[i].strokeColor = WHITE
            pie.slices[i].fontSize    = 7
            pie.slices[i].fontName    = "Helvetica-Bold"
        pie_d.add(pie)
        story.append(pie_d)
    story.append(Spacer(1, 12))

    if has_golden:
        story.append(sec_bar("Pass / Fail Summary by ID",
                              f"IoU threshold: {records[0]['iou_threshold']}"))
        story.append(Spacer(1, 6))
        hdr = ["ID", "Images", "PASS", "FAIL", "Pass Rate", "Avg IoU"]
        rows = []
        for prefix, stats in cat_data:
            rate = f"{round(stats['pass']/stats['count']*100,1)}%" if stats['count'] else "??"
            avg_i = f"{round(stats['iou_sum']/stats['iou_count'],3):.3f}" if stats['iou_count'] else "??"
            rows.append([prefix, stats["count"], stats["pass"], stats["fail"], rate, avg_i])
        cws = [PW*f for f in [0.28, 0.12, 0.12, 0.12, 0.18, 0.18]]
        t = mk_table(hdr, rows, cws)
        for ri, row in enumerate(rows, 1):
            t.setStyle(TableStyle([
                ("TEXTCOLOR", (2,ri),(2,ri), PASS),
                ("FONTNAME",  (2,ri),(2,ri), "Helvetica-Bold"),
                ("TEXTCOLOR", (3,ri),(3,ri), FAIL),
                ("FONTNAME",  (3,ri),(3,ri), "Helvetica-Bold"),
            ]))
        story.append(t)
        story.append(Spacer(1, 12))

    story.append(sec_bar("Component Class Breakdown"))
    story.append(Spacer(1, 6))
    grand  = sum(cls_vals) or 1
    c_rows = []
    for cls, total in sorted_classes:
        img_c = class_img_count[cls]
        avg   = round(total/img_c, 2) if img_c else 0
        pct   = f"{round(total/grand*100,1)}%"
        c_rows.append([cls, total, img_c, avg, pct])
    cws2 = [PW*f for f in [0.36, 0.16, 0.18, 0.15, 0.15]]
    story.append(mk_table(["Class","Total","Images","Avg/Image","Share"], c_rows, cws2))
    story.append(Spacer(1, 12))

    story.append(PageBreak())
    story.append(sec_bar(f"Raw Detection Data - first 60 rows of {len(records)}"))
    story.append(Spacer(1, 6))

    if has_golden:
        raw_hdr  = ["Image Name","Status","Matched","Avg IoU","ID","Mode"]
        raw_rows = []
        for r in records[:60]:
            matched_s = f"{r['matched']}/{r['total']}" if r['matched'] is not None else "??"
            iou_s     = f"{r['avg_iou']:.3f}" if r['avg_iou'] is not None else "??"
            raw_rows.append([r["image_name"], r["status"], matched_s, iou_s,
                             r["prefix"], r["golden_mode"]])
        cws3 = [PW*f for f in [0.30, 0.10, 0.13, 0.12, 0.20, 0.15]]
        story.append(mk_table(raw_hdr, raw_rows, cws3, status_col=1))
    else:
        raw_hdr  = ["Image Name","ID","Detected Classes","Components"]
        raw_rows = []
        for r in records[:60]:
            dc = r["detected_classes"][:60] + ("..." if len(r["detected_classes"]) > 60 else "")
            raw_rows.append([r["image_name"], r["prefix"], dc, r["total_components"]])
        cws3 = [PW*f for f in [0.22, 0.14, 0.52, 0.12]]
        story.append(mk_table(raw_hdr, raw_rows, cws3))

    if len(records) > 60:
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"...and {len(records)-60} more rows in the Excel report.",
                                body_style))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(GRAY)
        canvas.drawString(1.4*cm, 0.7*cm, "Component Detection Report - Auto-generated")
        canvas.drawRightString(W-1.4*cm, 0.7*cm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    print(f"  PDF   -> {out_path}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python detection_report_generator.py <input.csv or input.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"File not found: {path}")
        sys.exit(1)

    base     = os.path.splitext(path)[0]
    xlsx_out = base + "_report.xlsx"
    html_out = base + "_dashboard.html"
    pdf_out  = base + "_dashboard.pdf"

    print(f"\nReading: {path}")
    records, has_golden = load_data(path)
    fmt = "Format A (golden comparison)" if has_golden else "Format B (detection only)"
    print(f"   {len(records)} rows -> {fmt}")

    STATUS_ORDER = {"FAIL": 0, "PASS": 1}
    records.sort(key=lambda r: STATUS_ORDER.get(r["status"].upper(), 2))

    sc, ci, ps, st, iou_vals = aggregate(records)

    print("\nBuilding Excel...")
    build_excel(records, sc, ci, ps, st, iou_vals, has_golden, xlsx_out)
    print("Building HTML...")
    build_html(records, sc, ci, ps, st, iou_vals, has_golden, html_out)
    print("Building PDF...")
    build_pdf(records, sc, ci, ps, st, iou_vals, has_golden, pdf_out)

    print(f"\nDone.\n   {xlsx_out}\n   {html_out}\n   {pdf_out}\n")


if __name__ == "__main__":
    main()

